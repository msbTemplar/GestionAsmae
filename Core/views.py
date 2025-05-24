from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone
from django.urls import reverse
from .models import *
from .forms import *
import mimetypes
from django.core.paginator import Paginator
from django.db.models import Sum
from django.http import JsonResponse
from calendar import HTMLCalendar
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from django.utils.dateparse import parse_date

def liste_des_formulaire_charges_avec_filtre_Date_facture(request):
    factures = []
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    total_montant = 0
    chiffre_affaire = None
    difference = None
    
    if date_debut and date_fin:
        date_debut = parse_date(date_debut)
        date_fin = parse_date(date_fin)
        factures = FormulaireCharge.objects.filter(date_facture__range=(date_debut, date_fin)).select_related('charge')
        total_montant = factures.aggregate(total=Sum('montant_charge'))['total']
        
        # Obtener el año del rango (por simplicidad, asumimos que inicio y fin están en el mismo año)
        annee = date_debut.year
        try:
            chiffre_affaire = ChiffreAffaire.objects.get(annee=annee)
            difference = chiffre_affaire.montant - total_montant
        except ChiffreAffaire.DoesNotExist:
            chiffre_affaire = None
            difference = None
        
        if request.GET.get('export') == 'excel':
            # Crear un libro de trabajo y una hoja
            today = datetime.today()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Formulaire Charge"
            
            # Crear una nueva hoja para la tabla
            #ws2 = wb.create_sheet(title="Formulaire Concierge")

            # Agregar los encabezados de la tabla
            headers = ['Date Paiement', 'Charge', 'Date Facture', 'Mois', 'Montant', 'Nom du fichier']
            
            #if request.user.is_superuser:
                #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
            ws.append(headers)

            # Aplicar negrita a los encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Agregar los datos de la tabla
            for item in factures:
                row = [
                    item.date_payement.strftime('%Y-%m-%d %H:%M'),
                    item.charge.nome_charge if item.charge else '',
                    item.date_facture.strftime('%Y-%m-%d'),
                    item.get_mois_display(),  # Para mostrar el nombre del mes en lugar de '01', '02'...
                    float(item.montant_charge),
                    item.image_charge.name if item.image_charge else ''
                ]
                #if request.user.is_superuser:
                    #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
                ws.append(row)
                
            ws.append(['Total Montant', '', '','', total_montant])
            
            # Crear una respuesta HTTP con el archivo Excel
            filename = f"liste_charges_immeuble_filtrer_date_facture_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            return response
    
    return render(request, './liste_factures_filtrees.html', {
        'factures': factures,
        'date_debut': request.GET.get('date_debut', ''),
        'date_fin': request.GET.get('date_fin', ''),
        'total_montant': total_montant,
        'chiffre_affaire': chiffre_affaire,
        'difference': difference,
    })

def enregistrer_formulaire_charge_view(request):
    if request.method == 'POST':
        form = EnregistrerFormulaireChargeForm(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save(commit=False)  # Aún no lo guardamos del todo

            # Guardar la instancia primero para que `instance.image_charge.path` esté disponible
            instance.save()

            # Ahora los datos están disponibles
            charge = instance.charge
            date = instance.date_payement
            montant = instance.montant_charge
            image_charge = instance.image_charge

            # Crear el mensaje de correo
            email_message = EmailMessage(
                subject=f'Formulaire Charge: {date} - {charge} - {montant}',
                body=(
                    f'Nom de la charge : {charge}\n'
                    f'Date de la charge : {date}\n'
                    f'Montant de la charge : {montant}\n'
                
                ),
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )

            # Adjuntar archivo si hay
            if image_charge:
                mime_type, _ = mimetypes.guess_type(image_charge.path)
                with open(image_charge.path, 'rb') as f:
                    email_message.attach(image_charge.name, f.read(), mime_type)

            # Enviar email
            email_message.send(fail_silently=False)

            return redirect('liste_des_formulaire_charges')
    else:
        form = EnregistrerFormulaireChargeForm()

    return render(request, './enregistrer_formulaire_charge.html', {'form': form})

def actualiser_formulaire_charge(request, id_formulaire_charge):
    formulaire_charge = FormulaireCharge.objects.get(pk=id_formulaire_charge)
    form = EnregistrerFormulaireChargeForm(request.POST or None, request.FILES or None,  instance=formulaire_charge)
    if form.is_valid():
        form.save()
        return redirect('liste_des_formulaire_charges')
    context = {'formulaire_charge': formulaire_charge, 'form': form}
    return render(request, './actualizer_formulaire_charge.html', context)

def eliminer_formulaire_charge(request, id_formulaire_charge):
    formulaire_charge = get_object_or_404(FormulaireCharge, id=id_formulaire_charge)
    formulaire_charge.delete()
    messages.success(request, "Le formulaire charge a été eliminer correctement.")
    return redirect('liste_des_formulaire_charges')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

def liste_des_formulaire_charges(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_formulaire_charges = FormulaireCharge.objects.all()
    total_montant = la_lista_des_formulaire_charges.aggregate(total=Sum('montant_charge'))['total']
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_formulaire_charges, 5)
    page = request.GET.get('page')
    tous_les_formulaire_charges = p.get_page(page)
    nums = "a" * tous_les_formulaire_charges.paginator.num_pages
    
    print("hola : " + str(tous_les_formulaire_charges.paginator.num_pages))
    
    context = {'la_lista_des_formulaire_charges': la_lista_des_formulaire_charges, 'tous_les_formulaire_charges': tous_les_formulaire_charges, 'nums': nums, 'name':name, 'total_montant':total_montant}
    
    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        today = datetime.today()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulaire Charge"
        
        # Crear una nueva hoja para la tabla
        #ws2 = wb.create_sheet(title="Formulaire Concierge")

        # Agregar los encabezados de la tabla
        headers = ['Date Paiement', 'Charge', 'Date Facture', 'Mois', 'Montant', 'Nom du fichier']
        
        #if request.user.is_superuser:
            #headers.extend(['Imprimer pdf', 'Actualiser', 'Eliminer'])
        ws.append(headers)

        # Aplicar negrita a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar los datos de la tabla
        for item in la_lista_des_formulaire_charges:
            row = [
                item.date_payement.strftime('%Y-%m-%d %H:%M'),
                item.charge.nome_charge if item.charge else '',
                item.date_facture.strftime('%Y-%m-%d'),
                item.get_mois_display(),  # Para mostrar el nombre del mes en lugar de '01', '02'...
                float(item.montant_charge),
                item.image_charge.name if item.image_charge else ''
            ]
            #if request.user.is_superuser:
                #row.extend(['Imprimer', 'Actualiser', 'Eliminer'])
            ws.append(row)
            
        ws.append(['Total Montant', '', '','', total_montant])
        
        # Crear una respuesta HTTP con el archivo Excel
        filename = f"liste_charges_immeuble_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] = 'attachment; filename=lista_paie_concierge.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    return render(request, './la_liste_des_formulaire_charges.html', context)


def actualiser_la_charge(request, id_charge):
    charge = Charge.objects.get(pk=id_charge)
    form = EnregistrerChargeForm(request.POST or None, request.FILES or None,  instance=charge)
    if form.is_valid():
        form.save()
        return redirect('liste_des_charges')
    context = {'charge': charge, 'form': form}
    return render(request, './actualizer_la_charge.html', context)

def eliminer_la_charge(request, id_charge):
    charge = get_object_or_404(Charge, id=id_charge)
    charge.delete()
    messages.success(request, "La charge a été eliminer correctement.")
    return redirect('liste_des_charges')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal


def enregistrer_charge_view(request):
    if request.method == 'POST':
        form = EnregistrerChargeForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            nome_charge = instance.nome_charge
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {nome_charge} - {nome_charge}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Nom de la charge: {nome_charge}\nNom de la charge: {nome_charge}\nNom de la charge: {nome_charge}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if serie_pelicula_imagen:
                #mime_type, _ = mimetypes.guess_type(serie_pelicula_imagen.path)
                #email_message.attach(serie_pelicula_imagen.name, serie_pelicula_imagen.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_charges')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerChargeForm()
    return render(request, 'enregistrer_charge.html', {'form': form})


def liste_des_charges(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_lista_des_charges = Charge.objects.all()
    name = request.user.username
    # set pagination
    
    p = Paginator(la_lista_des_charges, 5)
    page = request.GET.get('page')
    tous_les_charges = p.get_page(page)
    nums = "a" * tous_les_charges.paginator.num_pages
    
    print("hola : " + str(tous_les_charges.paginator.num_pages))
    
    context = {'la_lista_des_charges': la_lista_des_charges, 'tous_les_charges': tous_les_charges, 'nums': nums, 'name':name}
    return render(request, './la_liste_des_charges.html', context)

def liste_des_chiffres_affaires(request):
    # venue_list = Venue.objects.all().order_by('?')
    la_liste_des_chiffres_affaires = ChiffreAffaire.objects.all()
    name = request.user.username
    # set pagination
    
    p = Paginator(la_liste_des_chiffres_affaires, 5)
    page = request.GET.get('page')
    tous_les_chiffres_affaires = p.get_page(page)
    nums = "a" * tous_les_chiffres_affaires.paginator.num_pages
    
    print("hola : " + str(tous_les_chiffres_affaires.paginator.num_pages))
    
    context = {'la_liste_des_chiffres_affaires': la_liste_des_chiffres_affaires, 'tous_les_chiffres_affaires': tous_les_chiffres_affaires, 'nums': nums, 'name':name}
    return render(request, './la_liste_des_chiffres_affaires.html', context)

def enregistrer_le_chiffre_affaire_view(request):
    if request.method == 'POST':
        form = EnregistrerChiffreAffaireForm(request.POST, request.FILES)
        if form.is_valid():
            #form.save()
            #titulo_serie = request.POST['titulo_serie']
            #serie_o_pelicula = request.POST['serie_o_pelicula']
            #plataforma = request.POST['plataforma']
            #name = request.user.username
            #file = request.FILES['file']
            instance = form.save()
            annee = instance.annee
            montant = instance.montant
           
            #files = request.FILES.getlist('files')
            email_message = EmailMessage(
                subject=f'Contact Form: {annee} - {montant}',
                #body=titulo_serie + " " + serie_o_pelicula + " " +  plataforma,
                body=f'Chiffre d affaire: {annee}\nMontant: {montant}',
                
                from_email=settings.EMAIL_HOST_USER,
                to=['msb.duck@gmail.com', 'msb.tesla@gmail.com', 'msebti2@gmail.com', 'msb.acer@gmail.com'],
                reply_to=['msebti2@gmail.com']
            )
            # Adjuntar cada archivo
            #for file in files:
                #email_message.attach(file.name, file.read(), file.content_type)
            
            #if serie_pelicula_imagen:
                #mime_type, _ = mimetypes.guess_type(serie_pelicula_imagen.path)
                #email_message.attach(serie_pelicula_imagen.name, serie_pelicula_imagen.read(), mime_type)
            
            # Adjuntar el archivo
            #email_message.attach(file.name, file.read(), file.content_type)

            # Enviar el email
            email_message.send(fail_silently=False)
            form.save()
            return redirect('liste_des_chiffres_affaires')  # Cambia esto por la vista a la que deseas redirigir después de guardar
    else:
        form = EnregistrerChiffreAffaireForm()
    return render(request, 'enregistrer_le_chiffre_affaire.html', {'form': form})

def actualizer_le_chiffre_affaire(request, id_chiffre_affaire):
    le_chiffre_affaire = ChiffreAffaire.objects.get(pk=id_chiffre_affaire)
    form = EnregistrerChiffreAffaireForm(request.POST or None, request.FILES or None,  instance=le_chiffre_affaire)
    if form.is_valid():
        form.save()
        return redirect('liste_des_chiffres_affaires')
    context = {'le_chiffre_affaire': le_chiffre_affaire, 'form': form}
    return render(request, './actualizer_le_chiffre_affaire.html', context)

def eliminer_le_chiffre_affaire(request, id_chiffre_affaire):
    le_chiffre_affaire = get_object_or_404(ChiffreAffaire, id=id_chiffre_affaire)
    le_chiffre_affaire.delete()
    messages.success(request, "Le chiffre de l'affaire a été eliminer correctement.")
    return redirect('liste_des_chiffres_affaires')  # Reemplaza 'nombre_de_tu_vista' con el nombre de tu vista principal

@login_required
def Home(request):
    return render(request, 'index.html')

def RegisterView(request):

    if request.method == "POST":
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')

        user_data_has_error = False

        if User.objects.filter(username=username).exists():
            user_data_has_error = True
            messages.error(request, "Username already exists")

        if User.objects.filter(email=email).exists():
            user_data_has_error = True
            messages.error(request, "Email already exists")

        if len(password) < 5:
            user_data_has_error = True
            messages.error(request, "Password must be at least 5 characters")

        if user_data_has_error:
            return redirect('register')
        else:
            new_user = User.objects.create_user(
                first_name=first_name,
                last_name=last_name,
                email=email, 
                username=username,
                password=password
            )
            messages.success(request, "Account created. Login now")
            return redirect('login')

    return render(request, 'register.html')

def LoginView(request):

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            return redirect('liste_des_formulaire_charges')
        
        else:
            messages.error(request, "Invalid login credentials")
            return redirect('login')

    return render(request, 'login.html')

def LogoutView(request):

    logout(request)

    return redirect('login')

def ForgotPassword(request):

    if request.method == "POST":
        email = request.POST.get('email')

        try:
            user = User.objects.get(email=email)

            new_password_reset = PasswordReset(user=user)
            new_password_reset.save()

            password_reset_url = reverse('reset-password', kwargs={'reset_id': new_password_reset.reset_id})

            full_password_reset_url = f'{request.scheme}://{request.get_host()}{password_reset_url}'

            email_body = f'Reset your password using the link below:\n\n\n{full_password_reset_url}'
        
            email_message = EmailMessage(
                'Reset your password', # email subject
                email_body,
                settings.EMAIL_HOST_USER, # email sender
                [email] # email  receiver 
            )

            email_message.fail_silently = True
            email_message.send()

            return redirect('password-reset-sent', reset_id=new_password_reset.reset_id)

        except User.DoesNotExist:
            messages.error(request, f"No user with email '{email}' found")
            return redirect('forgot-password')

    return render(request, 'forgot_password.html')

def PasswordResetSent(request, reset_id):

    if PasswordReset.objects.filter(reset_id=reset_id).exists():
        return render(request, 'password_reset_sent.html')
    else:
        # redirect to forgot password page if code does not exist
        messages.error(request, 'Invalid reset id')
        return redirect('forgot-password')

def ResetPassword(request, reset_id):

    try:
        password_reset_id = PasswordReset.objects.get(reset_id=reset_id)

        if request.method == "POST":
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')

            passwords_have_error = False

            if password != confirm_password:
                passwords_have_error = True
                messages.error(request, 'Passwords do not match')

            if len(password) < 5:
                passwords_have_error = True
                messages.error(request, 'Password must be at least 5 characters long')

            expiration_time = password_reset_id.created_when + timezone.timedelta(minutes=10)

            if timezone.now() > expiration_time:
                passwords_have_error = True
                messages.error(request, 'Reset link has expired')

                password_reset_id.delete()

            if not passwords_have_error:
                user = password_reset_id.user
                user.set_password(password)
                user.save()

                password_reset_id.delete()

                messages.success(request, 'Password reset. Proceed to login')
                return redirect('login')
            else:
                # redirect back to password reset page and display errors
                return redirect('reset-password', reset_id=reset_id)

    
    except PasswordReset.DoesNotExist:
        
        # redirect to forgot password page if code does not exist
        messages.error(request, 'Invalid reset id')
        return redirect('forgot-password')

    return render(request, 'reset_password.html')